"""
Exportações: Excel, PDF e HTML auto-contido (para partilha).

Todas com branding Núcleo Assets.
"""

from __future__ import annotations

import io
import html
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BRAND_NAME = "Núcleo Assets"
BRAND_COLOR = "1F4E78"
BRAND_ACCENT = "C8A951"


def _fmt_eur(v):
    if v is None or v != v:
        return "—"
    try:
        return f"{v:,.0f} €".replace(",", " ")
    except Exception:
        return "—"


def _fmt_pct(v):
    if v is None or v != v:
        return "—"
    try:
        return f"{v*100:.1f}%"
    except Exception:
        return "—"


# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

def export_excel(analise: dict, inputs_dict: dict) -> bytes:
    """Devolve bytes de um .xlsx com a análise completa, branding Núcleo Assets."""
    out = analise["outputs"]
    veredicto = analise["veredicto"]
    flags = analise.get("flags", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Análise"
    ws.sheet_view.showGridLines = False

    HEADER = PatternFill("solid", start_color=BRAND_COLOR)
    ACCENT = PatternFill("solid", start_color=BRAND_ACCENT)
    HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    TFONT = Font(name="Arial", bold=True, size=18, color=BRAND_COLOR)
    SFONT = Font(name="Arial", italic=True, size=10, color="595959")
    BFONT = Font(name="Arial", bold=True, size=11)
    NFONT = Font(name="Arial", size=11)
    BORDER = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # Branding header
    ws["A1"] = f"{BRAND_NAME} — Análise Fix & Flip"
    ws["A1"].font = TFONT
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Deal: {inputs_dict.get('nome_deal', '—')}   •   {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SFONT
    ws.merge_cells("A2:E2")

    # Veredicto
    ws["A4"] = "VEREDICTO"
    ws["A4"].font = HFONT
    ws["A4"].fill = HEADER
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A4:E4")
    ws.row_dimensions[4].height = 24

    ws["A5"] = veredicto.get("rotulo", "")
    ws["A5"].font = Font(name="Arial", bold=True, size=14, color=BRAND_COLOR)
    ws.merge_cells("A5:E5")
    ws["A6"] = veredicto.get("mensagem", "")
    ws["A6"].font = NFONT
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A6:E6")
    ws.row_dimensions[6].height = 40

    # Métricas-chave
    row = 8
    ws.cell(row=row, column=1, value="MÉTRICAS-CHAVE").font = HFONT
    ws.cell(row=row, column=1).fill = HEADER
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 22

    metrics = [
        ("Lucro líquido", _fmt_eur(out.get("lucro_liquido"))),
        ("ROI total", _fmt_pct(out.get("roi_total"))),
        ("Cash-on-Cash anual", _fmt_pct(out.get("cash_on_cash_anual"))),
        ("IRR aproximado", _fmt_pct(out.get("irr_aproximado"))),
        ("Equity total", _fmt_eur(out.get("equity_total"))),
        ("Preço/m² compra", _fmt_eur(out.get("preco_m2_compra"))),
        ("Preço/m² venda", _fmt_eur(out.get("preco_m2_venda"))),
        ("Preço venda usado", _fmt_eur(out.get("preco_venda_usado"))),
        ("Margem bruta (% venda)", _fmt_pct(out.get("margem_bruta_pct_venda"))),
    ]
    row += 1
    for label, val in metrics:
        ws.cell(row=row, column=1, value=label).font = BFONT
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=val).font = NFONT
        ws.cell(row=row, column=2).border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    # Inputs
    row += 1
    ws.cell(row=row, column=1, value="INPUTS").font = HFONT
    ws.cell(row=row, column=1).fill = HEADER
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 22
    row += 1

    input_keys = [
        ("nome_deal", "Nome do deal"),
        ("freguesia", "Freguesia"),
        ("tipo_imovel", "Tipo imóvel"),
        ("tipologia", "Tipologia"),
        ("area_m2", "Área (m²)"),
        ("estado_conservacao", "Estado"),
        ("preco_aquisicao", "Preço aquisição"),
        ("ciclo_meses", "Ciclo (meses)"),
        ("venda_modo", "Modo cálculo venda"),
        ("estrutura", "Estrutura fiscal"),
        ("usa_financiamento", "Usa financiamento"),
        ("ltv", "LTV"),
    ]
    for key, label in input_keys:
        v = inputs_dict.get(key)
        if isinstance(v, float) and key in ("ltv",):
            v_str = _fmt_pct(v)
        elif isinstance(v, (int, float)) and key in ("preco_aquisicao",):
            v_str = _fmt_eur(v)
        else:
            v_str = str(v) if v is not None else "—"
        ws.cell(row=row, column=1, value=label).font = BFONT
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=v_str).font = NFONT
        ws.cell(row=row, column=2).border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    # Breakdown
    row += 1
    ws.cell(row=row, column=1, value="BREAKDOWN DE CUSTOS").font = HFONT
    ws.cell(row=row, column=1).fill = HEADER
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.row_dimensions[row].height = 22
    row += 1
    breakdown = [
        ("IMT", out.get("imt")),
        ("IS (transmissão)", out.get("is_transmissao")),
        ("Notário+registo", inputs_dict.get("custos_notario_registo")),
        ("Custos compra total", out.get("custos_compra_total")),
        ("Obra (final com IVA + cont.)", out.get("obra_com_contingencia")),
        ("Juros ciclo (+comissões)", out.get("juros_pagos_ciclo")),
        ("IMI ciclo", out.get("imi_ciclo")),
        ("Holding (condomínio etc)", out.get("custos_holding_total")),
        ("Seguros ciclo", out.get("seguros_ciclo")),
        ("Custos venda (comissão)", out.get("custos_venda")),
        ("Imposto saída", out.get("imposto_saida")),
        ("LUCRO LÍQUIDO", out.get("lucro_liquido")),
        ("Equity total", out.get("equity_total")),
        ("Valor empréstimo", out.get("valor_emprestimo")),
    ]
    for label, val in breakdown:
        cell_lab = ws.cell(row=row, column=1, value=label)
        cell_lab.font = BFONT if "LUCRO" in label else NFONT
        cell_lab.border = BORDER
        cell_val = ws.cell(row=row, column=2, value=_fmt_eur(val))
        cell_val.font = BFONT if "LUCRO" in label else NFONT
        cell_val.border = BORDER
        if "LUCRO" in label:
            cell_lab.fill = ACCENT
            cell_val.fill = ACCENT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    # Flags
    if flags:
        row += 1
        ws.cell(row=row, column=1, value="FLAGS").font = HFONT
        ws.cell(row=row, column=1).fill = HEADER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 22
        row += 1
        for f in flags:
            nivel_label = {"ok": "OK", "amarelo": "ATENÇÃO", "vermelho": "RISCO"}.get(f["nivel"], f["nivel"])
            ws.cell(row=row, column=1, value=nivel_label).font = BFONT
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2, value=f["mensagem"]).font = NFONT
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=2).border = BORDER
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            ws.row_dimensions[row].height = 30
            row += 1

    # Rodapé
    row += 2
    ws.cell(row=row, column=1, value=f"Gerado pelo Avaliador de Propriedades · {BRAND_NAME}").font = SFONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    # Larguras
    ws.column_dimensions["A"].width = 32
    for c in ["B", "C", "D", "E"]:
        ws.column_dimensions[c].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# HTML auto-contido (partilha)
# ---------------------------------------------------------------------------

def export_html(analise: dict, inputs_dict: dict) -> str:
    """Devolve string HTML auto-contido (CSS inline) para partilhar/abrir."""
    out = analise["outputs"]
    veredicto = analise["veredicto"]
    flags = analise.get("flags", [])

    cor_map = {"verde": "#1F8A3B", "amarelo": "#C8A951", "vermelho": "#B33A3A", "ok": "#1F8A3B"}
    cor_veredicto = cor_map.get(veredicto.get("cor"), "#595959")

    def esc(x):
        return html.escape(str(x)) if x is not None else "—"

    metrics_rows = [
        ("Lucro líquido", _fmt_eur(out.get("lucro_liquido"))),
        ("ROI total", _fmt_pct(out.get("roi_total"))),
        ("Cash-on-Cash anual", _fmt_pct(out.get("cash_on_cash_anual"))),
        ("IRR aproximado", _fmt_pct(out.get("irr_aproximado"))),
        ("Equity total", _fmt_eur(out.get("equity_total"))),
        ("Preço/m² compra", _fmt_eur(out.get("preco_m2_compra"))),
        ("Preço/m² venda", _fmt_eur(out.get("preco_m2_venda"))),
        ("Preço venda usado", _fmt_eur(out.get("preco_venda_usado"))),
        ("Margem bruta (% venda)", _fmt_pct(out.get("margem_bruta_pct_venda"))),
    ]

    breakdown_rows = [
        ("IMT", out.get("imt")),
        ("IS (transmissão)", out.get("is_transmissao")),
        ("Notário+registo", inputs_dict.get("custos_notario_registo")),
        ("Custos compra total", out.get("custos_compra_total")),
        ("Obra (final c/ IVA + cont.)", out.get("obra_com_contingencia")),
        ("Juros ciclo", out.get("juros_pagos_ciclo")),
        ("IMI ciclo", out.get("imi_ciclo")),
        ("Holding", out.get("custos_holding_total")),
        ("Seguros ciclo", out.get("seguros_ciclo")),
        ("Custos venda", out.get("custos_venda")),
        ("Imposto saída", out.get("imposto_saida")),
        ("Lucro líquido", out.get("lucro_liquido")),
    ]

    metric_cells = "".join(
        f'<div class="metric"><div class="m-label">{esc(lab)}</div>'
        f'<div class="m-value">{esc(val)}</div></div>'
        for lab, val in metrics_rows
    )

    breakdown_rows_html = "".join(
        f'<tr class="{"highlight" if "Lucro" in lab else ""}">'
        f'<td>{esc(lab)}</td><td class="num">{_fmt_eur(val)}</td></tr>'
        for lab, val in breakdown_rows
    )

    flags_html = ""
    if flags:
        flag_items = "".join(
            f'<li class="flag-{esc(f["nivel"])}">{esc(f["mensagem"])}</li>'
            for f in flags
        )
        flags_html = f'<section><h2>Flags</h2><ul class="flags">{flag_items}</ul></section>'

    inputs_html = "".join(
        f'<tr><td>{esc(lab)}</td><td>{esc(inputs_dict.get(key, "—"))}</td></tr>'
        for key, lab in [
            ("nome_deal", "Deal"),
            ("freguesia", "Freguesia"),
            ("tipo_imovel", "Tipo imóvel"),
            ("tipologia", "Tipologia"),
            ("area_m2", "Área (m²)"),
            ("estado_conservacao", "Estado"),
            ("preco_aquisicao", "Preço aquisição"),
            ("ciclo_meses", "Ciclo (meses)"),
            ("venda_modo", "Modo venda"),
            ("estrutura", "Estrutura fiscal"),
        ]
    )

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    return f"""<!DOCTYPE html>
<html lang="pt-PT">
<head>
<meta charset="utf-8">
<title>{esc(inputs_dict.get('nome_deal', 'Análise'))} — {BRAND_NAME}</title>
<style>
* {{ box-sizing: border-box; }}
body {{
    font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif;
    max-width: 960px; margin: 0 auto; padding: 32px 24px;
    color: #1a1a1a; background: #fafafa; line-height: 1.5;
}}
header {{
    border-bottom: 3px solid #{BRAND_COLOR}; padding-bottom: 16px; margin-bottom: 24px;
}}
.brand {{ color: #{BRAND_COLOR}; font-weight: 700; font-size: 14px; letter-spacing: 1px; text-transform: uppercase; }}
h1 {{ margin: 4px 0 0; font-size: 28px; color: #1a1a1a; }}
.sub {{ color: #666; font-size: 13px; margin-top: 4px; }}
.veredicto {{
    padding: 16px 20px; border-radius: 8px; margin-bottom: 24px;
    background: {cor_veredicto}; color: white;
}}
.veredicto h2 {{ margin: 0 0 6px; font-size: 20px; }}
.veredicto p {{ margin: 0; font-size: 14px; opacity: 0.95; }}
section {{ background: white; border-radius: 8px; padding: 20px 24px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }}
section h2 {{ margin: 0 0 16px; font-size: 16px; color: #{BRAND_COLOR}; text-transform: uppercase; letter-spacing: 0.5px; }}
.metrics {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; }}
.metric {{ background: #f5f7fa; border-left: 3px solid #{BRAND_COLOR}; padding: 12px 14px; border-radius: 4px; }}
.m-label {{ font-size: 11px; color: #666; text-transform: uppercase; letter-spacing: 0.5px; }}
.m-value {{ font-size: 18px; font-weight: 600; margin-top: 4px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
td {{ padding: 8px 4px; border-bottom: 1px solid #eee; }}
td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
tr.highlight td {{ font-weight: 700; background: #fff8e1; }}
ul.flags {{ list-style: none; padding: 0; margin: 0; }}
ul.flags li {{ padding: 8px 12px; border-radius: 4px; margin-bottom: 6px; font-size: 14px; }}
li.flag-ok {{ background: #e6f4ea; color: #1F8A3B; }}
li.flag-amarelo {{ background: #fdf6e3; color: #8a6d00; }}
li.flag-vermelho {{ background: #fde2e2; color: #B33A3A; }}
footer {{ text-align: center; color: #888; font-size: 12px; margin-top: 32px; }}
@media print {{ body {{ background: white; }} section {{ box-shadow: none; border: 1px solid #ddd; }} }}
</style>
</head>
<body>
<header>
    <div class="brand">{BRAND_NAME}</div>
    <h1>{esc(inputs_dict.get('nome_deal', 'Análise Fix & Flip'))}</h1>
    <div class="sub">{esc(inputs_dict.get('freguesia', ''))} · {esc(inputs_dict.get('tipologia', ''))} · {esc(inputs_dict.get('area_m2', ''))} m² · {timestamp}</div>
</header>

<div class="veredicto">
    <h2>{esc(veredicto.get('rotulo', ''))}</h2>
    <p>{esc(veredicto.get('mensagem', ''))}</p>
</div>

<section>
    <h2>Métricas-chave</h2>
    <div class="metrics">{metric_cells}</div>
</section>

<section>
    <h2>Inputs principais</h2>
    <table>{inputs_html}</table>
</section>

<section>
    <h2>Breakdown</h2>
    <table>{breakdown_rows_html}</table>
</section>

{flags_html}

<footer>Gerado pelo Avaliador de Propriedades · {BRAND_NAME} · {timestamp}</footer>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# PDF — abordagem pragmática: HTML → PDF via reportlab (texto puro estilizado)
# Para evitar dependência de wkhtmltopdf, usamos reportlab directamente.
# ---------------------------------------------------------------------------

def export_pdf(analise: dict, inputs_dict: dict) -> bytes:
    """Gera PDF com branding Núcleo Assets usando reportlab."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
        )
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError:
        raise RuntimeError(
            "reportlab não está instalado. Adiciona 'reportlab' ao requirements.txt "
            "e corre: pip install reportlab"
        )

    out = analise["outputs"]
    veredicto = analise["veredicto"]
    flags = analise.get("flags", [])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    brand_col = HexColor(f"#{BRAND_COLOR}")
    accent_col = HexColor(f"#{BRAND_ACCENT}")
    cor_map = {"verde": "#1F8A3B", "amarelo": "#C8A951", "vermelho": "#B33A3A", "ok": "#1F8A3B"}
    veredicto_col = HexColor(cor_map.get(veredicto.get("cor"), "#595959"))

    styles = getSampleStyleSheet()
    brand_style = ParagraphStyle(
        "Brand", parent=styles["Normal"], fontSize=10, textColor=brand_col,
        spaceAfter=2, fontName="Helvetica-Bold",
    )
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontSize=22, textColor=HexColor("#1a1a1a"),
        spaceAfter=4, alignment=TA_LEFT, fontName="Helvetica-Bold",
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=10, textColor=HexColor("#666666"),
        spaceAfter=18,
    )
    h2_style = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontSize=12, textColor=brand_col,
        spaceBefore=12, spaceAfter=8, fontName="Helvetica-Bold",
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"], fontSize=10, textColor=HexColor("#1a1a1a"),
    )
    veredicto_style = ParagraphStyle(
        "Veredicto", parent=styles["Normal"], fontSize=14, textColor=veredicto_col,
        fontName="Helvetica-Bold", spaceAfter=4,
    )

    story = []

    # Header
    story.append(Paragraph(BRAND_NAME.upper(), brand_style))
    story.append(Paragraph(html.escape(inputs_dict.get("nome_deal", "Análise Fix & Flip")), title_style))
    sub = f"{inputs_dict.get('freguesia') or '—'} · {inputs_dict.get('tipologia') or '—'} · {inputs_dict.get('area_m2') or '—'} m² · {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    story.append(Paragraph(html.escape(sub), sub_style))

    # Veredicto
    story.append(Paragraph("VEREDICTO", h2_style))
    story.append(Paragraph(html.escape(veredicto.get("rotulo", "")), veredicto_style))
    story.append(Paragraph(html.escape(veredicto.get("mensagem", "")), body_style))

    # Métricas
    story.append(Paragraph("MÉTRICAS-CHAVE", h2_style))
    metrics_data = [
        ["Lucro líquido", _fmt_eur(out.get("lucro_liquido")), "ROI total", _fmt_pct(out.get("roi_total"))],
        ["Cash-on-Cash anual", _fmt_pct(out.get("cash_on_cash_anual")), "IRR aproximado", _fmt_pct(out.get("irr_aproximado"))],
        ["Equity total", _fmt_eur(out.get("equity_total")), "Margem bruta", _fmt_pct(out.get("margem_bruta_pct_venda"))],
        ["Preço/m² compra", _fmt_eur(out.get("preco_m2_compra")), "Preço/m² venda", _fmt_eur(out.get("preco_m2_venda"))],
    ]
    t = Table(metrics_data, colWidths=[4.5*cm, 3.5*cm, 4.5*cm, 3.5*cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), HexColor("#666666")),
        ("TEXTCOLOR", (2, 0), (2, -1), HexColor("#666666")),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, HexColor("#dddddd")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(t)

    # Inputs principais
    story.append(Paragraph("INPUTS PRINCIPAIS", h2_style))
    input_data = [
        ["Tipo imóvel", str(inputs_dict.get("tipo_imovel") or "—")],
        ["Estado conservação", str(inputs_dict.get("estado_conservacao") or "—")],
        ["Preço aquisição", _fmt_eur(inputs_dict.get("preco_aquisicao"))],
        ["Ciclo (meses)", str(inputs_dict.get("ciclo_meses") or "—")],
        ["Modo cálculo venda", str(inputs_dict.get("venda_modo") or "—")],
        ["Estrutura fiscal", str(inputs_dict.get("estrutura") or "—")],
        ["Usa financiamento", "Sim" if inputs_dict.get("usa_financiamento") else "Não"],
    ]
    t2 = Table(input_data, colWidths=[6*cm, 10*cm])
    t2.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, HexColor("#eeeeee")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t2)

    # Breakdown
    story.append(Paragraph("BREAKDOWN DE CUSTOS", h2_style))
    breakdown = [
        ("IMT", out.get("imt")),
        ("IS (transmissão)", out.get("is_transmissao")),
        ("Notário+registo", inputs_dict.get("custos_notario_registo")),
        ("Custos compra total", out.get("custos_compra_total")),
        ("Obra (final c/ IVA + cont.)", out.get("obra_com_contingencia")),
        ("Juros ciclo", out.get("juros_pagos_ciclo")),
        ("IMI ciclo", out.get("imi_ciclo")),
        ("Holding", out.get("custos_holding_total")),
        ("Seguros ciclo", out.get("seguros_ciclo")),
        ("Custos venda", out.get("custos_venda")),
        ("Imposto saída", out.get("imposto_saida")),
        ("LUCRO LÍQUIDO", out.get("lucro_liquido")),
    ]
    breakdown_data = [[lab, _fmt_eur(val)] for lab, val in breakdown]
    t3 = Table(breakdown_data, colWidths=[10*cm, 6*cm])
    t3.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, HexColor("#eeeeee")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), HexColor("#FFF8E1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t3)

    # Flags
    if flags:
        story.append(Paragraph("FLAGS", h2_style))
        flag_style_map = {
            "ok": HexColor("#1F8A3B"),
            "amarelo": HexColor("#8a6d00"),
            "vermelho": HexColor("#B33A3A"),
        }
        for f in flags:
            fs = ParagraphStyle(
                "Flag", parent=body_style, fontSize=10,
                textColor=flag_style_map.get(f["nivel"], HexColor("#1a1a1a")),
                leftIndent=10, spaceAfter=4,
            )
            prefix = {"ok": "[OK]", "amarelo": "[ATENÇÃO]", "vermelho": "[RISCO]"}.get(f["nivel"], "[•]")
            story.append(Paragraph(f"{prefix} {html.escape(f['mensagem'])}", fs))

    # Footer
    story.append(Spacer(1, 0.8*cm))
    footer_style = ParagraphStyle(
        "Footer", parent=styles["Normal"], fontSize=8, textColor=HexColor("#888888"),
        alignment=TA_CENTER,
    )
    story.append(Paragraph(
        f"Gerado pelo Avaliador de Propriedades · {BRAND_NAME} · {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        footer_style,
    ))

    doc.build(story)
    return buf.getvalue()
